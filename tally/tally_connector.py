import requests
import json
import xml.etree.ElementTree as ET
import re
from datetime import datetime, date
from logging_config import logger
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

class TallyConnector:
    def __init__(self, host='localhost', port=9000):
        self.host = host
        self.port = port
        self.url = f'http://{host}:{port}'
        self.header = {'Content-Type': 'text/xml; charset=utf-8'}
        self.status = 'Disconnected'
        
        logger.info(f'Initializing TallyConnector for {self.url}')
        self.connect()

    def connect(self):
        try:
            logger.info(f'Attempting to connect to Tally at {self.url}...')
            response = requests.post(url=self.url, headers=self.header, timeout=60)
            
            if response.status_code == 200:
                self.status = 'Connected'
                logger.info(f'Successfully connected to Tally at {self.url}')
            else:
                self.status = 'Disconnected'
                logger.warning(
                    f'Unexpected response from Tally at {self.url}. '
                    f'Status code: {response.status_code}, Response: {response.text}'
                )

        except requests.exceptions.ConnectionError as e:
            self.status = 'Disconnected'
            logger.error(f'ConnectionError while connecting to Tally at {self.url}: {e}', exc_info=True)
        except requests.exceptions.Timeout as e:
            self.status = 'Disconnected'
            logger.error(f'Timeout while connecting to Tally at {self.url}: {e}', exc_info=True)
        except Exception as e:
            self.status = 'Disconnected'
            logger.error(f'Unexpected error while connecting to Tally at {self.url}: {e}', exc_info=True)

    def get_company_list(self):
        try:
            tree = ET.parse('utils/company.xml')
            xml_payload = ET.tostring(tree.getroot(), encoding='utf-8')
            response = requests.post(url=self.url, headers=self.header, data=xml_payload, timeout=120)

            if response.status_code == 200:
                root = ET.fromstring(response.content)
                all_companies = []

                for company in root.findall(".//COMPANY"):
                    name = company.find("NAME").text if company.find("NAME") is not None else "N/A"
                    start = company.find("STARTINGFROM").text if company.find("STARTINGFROM") is not None else "N/A"
                    gst = company.find("GSTIN").text if company.find("GSTIN") is not None else "N/A"

                    if name != "N/A" and name.strip() != "":
                        all_companies.append({
                            "name": name,
                            "start_from": start,
                            "gstin": gst,
                        })
                
                logger.info(f'Found {len(all_companies)} companies')
                return all_companies
            return []
        except Exception as e:
            logger.error(f'Error fetching company list: {e}', exc_info=True)
            return []

    def _get_text(self, element, tag, default=""):
        """Helper method to safely extract text from XML element"""
        elem = element.find(tag)
        return elem.text.strip() if elem is not None and elem.text else default

    def get_ledger(self, company_name):
        try:
            logger.info(f'Fetching ALL ledgers for company: {company_name}')
            
            all_ledgers = []
            
            tree = ET.parse('utils/ledger.xml')
            xml_payload = ET.tostring(tree.getroot(), encoding='utf-8')
            
            start_time = datetime.now()
            
            try:
                response = requests.post(
                    url=self.url, 
                    headers=self.header, 
                    data=xml_payload,
                    timeout=900,
                    stream=False
                )
            except requests.exceptions.Timeout:
                logger.error(f'Timeout (900s) fetching ledgers for {company_name}')
                return []
            
            request_time = (datetime.now() - start_time).total_seconds()
            
            if response.status_code != 200:
                logger.error(f'Failed to fetch ledgers. Status: {response.status_code}')
                return []
            
            debug_file = f"tally_ledger_response_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xml"
            with open(debug_file, 'wb') as f:
                f.write(response.content)
            logger.info(f'Saved XML response to {debug_file}')
            
            try:
                content = response.content.decode('utf-8', errors='ignore')
            except Exception as e:
                logger.error(f'Error decoding response: {e}')
                return []
            
            content = re.sub(r'[\x00-\x08\x0B\x0C\x0E-\x1F\x7F]', '', content)
            content = re.sub(r'&(?!amp;|lt;|gt;|quot;|apos;)', '&amp;', content)
            
            try:
                root = ET.fromstring(content.encode('utf-8'))
            except ET.ParseError as e:
                logger.error(f'XML Parse Error: {e}')
                return []
            
            ledgers = root.findall(".//LEDGER")
            
            if not ledgers:
                logger.info(f'No ledgers found for company: {company_name}')
                return []
            
            total_ledgers = len(ledgers)
            
            for idx, ledger in enumerate(ledgers, 1):
                try:
                    # Extract all possible fields comprehensively
                    ledger_data = {
                        # Basic Information
                        "name": ledger.get('NAME', ''),
                        "parent": self._get_text(ledger, "PARENT"),
                        "guid": self._get_text(ledger, "GUID"),
                        "alias": self._get_text(ledger, "ALIAS"),
                        
                        # Balances
                        "opening_balance": self._get_text(ledger, "OPENINGBALANCE", "0"),
                        "closing_balance": self._get_text(ledger, "CLOSINGBALANCE", "0"),
                        
                        # Contact Details
                        "mailing_name": self._get_text(ledger, "MAILINGNAME"),
                        "mobile": self._get_text(ledger, "LEDGERPHONE") or self._get_text(ledger, "PHONE"),
                        "email": self._get_text(ledger, "EMAIL") or self._get_text(ledger, "LEDGEREMAIL"),
                        "contact_person": self._get_text(ledger, "CONTACTPERSON"),
                        
                        # Billing/Primary Address
                        "address_line1": self._get_text(ledger, "ADDRESS") or self._get_text(ledger, ".//ADDRESS[1]"),
                        "address_line2": self._get_text(ledger, ".//ADDRESS[2]"),
                        "address_line3": self._get_text(ledger, ".//ADDRESS[3]"),
                        "city": self._get_text(ledger, "CITY"),
                        "state": self._get_text(ledger, "LEDSTATENAME") or self._get_text(ledger, "STATENAME"),
                        "pincode": self._get_text(ledger, "PINCODE"),
                        "country": self._get_text(ledger, "COUNTRYNAME"),
                        
                        # Shipping Address
                        "ship_address_line1": "",
                        "ship_address_line2": "",
                        "ship_address_line3": "",
                        "ship_city": "",
                        "ship_state": "",
                        "ship_pincode": "",
                        "ship_country": "",
                        
                        # Tax Information
                        "gstin": self._get_text(ledger, "PARTYGSTIN") or self._get_text(ledger, "GSTIN"),
                        "pan": self._get_text(ledger, "INCOMETAXNUMBER") or self._get_text(ledger, "PAN"),
                        "gst_registration_type": self._get_text(ledger, "GSTREGISTRATIONTYPE"),
                        "tax_rate": self._get_text(ledger, "TAXRATE"),
                        
                        # Bank Details
                        "bank_name": self._get_text(ledger, "BANKNAME"),
                        "bank_account_holder": self._get_text(ledger, "BANKACCHOLDERNAME"),
                        "bank_account_number": self._get_text(ledger, "ACCOUNTNUMBER") or self._get_text(ledger, "BANKACCOUNTNUMBER"),
                        "ifsc_code": self._get_text(ledger, "IFSCODE") or self._get_text(ledger, "BANKIFSCCODE"),
                        "branch": self._get_text(ledger, "BRANCHNAME"),
                        "swift_code": self._get_text(ledger, "SWIFTCODE"),
                        
                        # Credit/Payment Terms
                        "credit_limit": self._get_text(ledger, "CREDITLIMIT", "0"),
                        "credit_period": self._get_text(ledger, "CREDITPERIOD"),
                        "bill_credit_period": self._get_text(ledger, "BILLCREDITPERIOD"),
                        
                        # Ledger Configuration
                        "is_revenue": self._get_text(ledger, "ISREVENUE"),
                        "is_deemedpositive": self._get_text(ledger, "ISDEEMEDPOSITIVE"),
                        "affect_gross_profit": self._get_text(ledger, "AFFECTGROSSPROFIT"),
                        "inventory_values_affected": self._get_text(ledger, "INVENTORYVALUESAREAFFECTED"),
                        
                        # Additional Fields
                        "ledger_code": self._get_text(ledger, "LEDGERCODE"),
                        "ledger_mobile": self._get_text(ledger, "LEDGERMOBILE"),
                        "ledger_contact": self._get_text(ledger, "LEDGERCONTACT"),
                        "website": self._get_text(ledger, "WEBSITE"),
                        "nature_of_business": self._get_text(ledger, "NATUREOFBUSINESS"),
                        
                        # Dates
                        "date_of_incorporation": self._get_text(ledger, "DATEOFINCORPORATION"),
                        "financial_year_from": self._get_text(ledger, "FINANCIALYEARFROM"),
                        "books_beginning_from": self._get_text(ledger, "BOOKSBEGINNINGFROM"),
                        
                        # Organization Details
                        "organization_type": self._get_text(ledger, "ORGANIZATIONTYPE"),
                        "company_category": self._get_text(ledger, "COMPANYCATEGORY"),
                        
                        # TDS/TCS
                        "tds_applicable": self._get_text(ledger, "TDSAPPLICABLE"),
                        "tcs_applicable": self._get_text(ledger, "TCSAPPLICABLE"),
                        
                        # Multi-currency
                        "currency": self._get_text(ledger, "CURRENCY"),
                        "currency_symbol": self._get_text(ledger, "CURRENCYSYMBOL"),
                        
                        # Other Important Fields
                        "alterid": self._get_text(ledger, "ALTERID"),
                        "reserve1": self._get_text(ledger, "RESERVE1"),
                        "reserve2": self._get_text(ledger, "RESERVE2"),
                        "custom_field1": self._get_text(ledger, "CUSTOMFIELD1"),
                        "custom_field2": self._get_text(ledger, "CUSTOMFIELD2"),
                    }
                    
                    # Extract shipping address separately
                    # Look for shipping address elements (common variations)
                    ship_address_elems = ledger.findall(".//SHIPPINGADDRESS") or \
                                       ledger.findall(".//DELIVERYADDRESS") or \
                                       ledger.findall(".//MAILINGADDRESS")
                    
                    if ship_address_elems:
                        ship_elem = ship_address_elems[0]
                        ship_addresses = ship_elem.findall(".//ADDRESS")
                        if len(ship_addresses) > 0:
                            ledger_data["ship_address_line1"] = ship_addresses[0].text.strip() if ship_addresses[0].text else ""
                        if len(ship_addresses) > 1:
                            ledger_data["ship_address_line2"] = ship_addresses[1].text.strip() if ship_addresses[1].text else ""
                        if len(ship_addresses) > 2:
                            ledger_data["ship_address_line3"] = ship_addresses[2].text.strip() if ship_addresses[2].text else ""
                        
                        ledger_data["ship_city"] = self._get_text(ship_elem, "CITY")
                        ledger_data["ship_state"] = self._get_text(ship_elem, "STATENAME")
                        ledger_data["ship_pincode"] = self._get_text(ship_elem, "PINCODE")
                        ledger_data["ship_country"] = self._get_text(ship_elem, "COUNTRYNAME")
                    
                    # Combine all address lines for billing
                    address_list = ledger.findall(".//ADDRESS")
                    if address_list and not ledger_data["address_line1"]:
                        if len(address_list) > 0:
                            ledger_data["address_line1"] = address_list[0].text.strip() if address_list[0].text else ""
                        if len(address_list) > 1:
                            ledger_data["address_line2"] = address_list[1].text.strip() if address_list[1].text else ""
                        if len(address_list) > 2:
                            ledger_data["address_line3"] = address_list[2].text.strip() if address_list[2].text else ""
                    
                    if ledger_data["name"]:
                        all_ledgers.append(ledger_data)
                
                except Exception as e:
                    logger.warning(f'Error processing ledger at index {idx}: {e}')
                    continue
            
            total_time = (datetime.now() - start_time).total_seconds()
            
            logger.info(f'Successfully fetched {len(all_ledgers)} ledgers for {company_name} in {total_time:.1f}s')
            
            return all_ledgers
            
        except requests.exceptions.ConnectionError as e:
            logger.error(f'Connection error for {company_name}: {e}', exc_info=True)
            return []
            
        except MemoryError as e:
            logger.error(f'Memory error processing {company_name}: {e}', exc_info=True)
            return []
            
        except Exception as e:
            logger.error(f'Unexpected error for {company_name}: {e}', exc_info=True)
            return []

    def save_ledgers_to_excel(self, ledgers, company_name=""):
        """Export ledgers to Excel with all fields and proper formatting"""
        try:
            if not ledgers:
                logger.warning('Attempted to export empty ledger list to Excel')
                return None
            
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            safe_company_name = "".join(c for c in company_name if c.isalnum() or c in (' ', '-', '_')).strip()
            safe_company_name = safe_company_name[:50]
            
            if safe_company_name:
                filename = f"ledgers_{safe_company_name}_{timestamp}.xlsx"
            else:
                filename = f"ledgers_{timestamp}.xlsx"
            
            # Create workbook
            wb = Workbook()
            sheet = wb.active
            sheet.title = "Ledgers"
            
            # Define all columns
            headers = [
                # Basic Information
                'Ledger Name', 'Parent Group', 'GUID', 'Alias',
                
                # Balances
                'Opening Balance', 'Closing Balance',
                
                # Contact Details
                'Mailing Name', 'Mobile', 'Email', 'Contact Person',
                
                # Billing/Primary Address
                'Address Line 1', 'Address Line 2', 'Address Line 3',
                'City', 'State', 'Pincode', 'Country',
                
                # Shipping Address
                'Ship Address Line 1', 'Ship Address Line 2', 'Ship Address Line 3',
                'Ship City', 'Ship State', 'Ship Pincode', 'Ship Country',
                
                # Tax Information
                'GSTIN', 'PAN', 'GST Registration Type', 'Tax Rate',
                
                # Bank Details
                'Bank Name', 'Bank Account Holder', 'Bank Account Number',
                'IFSC Code', 'Branch', 'Swift Code',
                
                # Credit/Payment Terms
                'Credit Limit', 'Credit Period', 'Bill Credit Period',
                
                # Ledger Configuration
                'Is Revenue', 'Is Deemed Positive', 'Affect Gross Profit', 
                'Inventory Values Affected',
                
                # Additional Fields
                'Ledger Code', 'Ledger Mobile', 'Ledger Contact', 'Website',
                'Nature of Business', 'Date of Incorporation', 'Financial Year From',
                'Books Beginning From', 'Organization Type', 'Company Category',
                
                # TDS/TCS
                'TDS Applicable', 'TCS Applicable',
                
                # Multi-currency
                'Currency', 'Currency Symbol',
                
                # Other
                'Alter ID', 'Reserve 1', 'Reserve 2', 'Custom Field 1', 'Custom Field 2'
            ]
            
            # Write headers with formatting
            for col_idx, header in enumerate(headers, 1):
                cell = sheet.cell(row=1, column=col_idx, value=header)
                cell.font = Font(bold=True, color='FFFFFF')
                cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            # Write data
            for row_idx, ledger in enumerate(ledgers, 2):
                data_row = [
                    ledger.get('name', ''),
                    ledger.get('parent', ''),
                    ledger.get('guid', ''),
                    ledger.get('alias', ''),
                    ledger.get('opening_balance', '0'),
                    ledger.get('closing_balance', '0'),
                    ledger.get('mailing_name', ''),
                    ledger.get('mobile', ''),
                    ledger.get('email', ''),
                    ledger.get('contact_person', ''),
                    ledger.get('address_line1', ''),
                    ledger.get('address_line2', ''),
                    ledger.get('address_line3', ''),
                    ledger.get('city', ''),
                    ledger.get('state', ''),
                    ledger.get('pincode', ''),
                    ledger.get('country', ''),
                    ledger.get('ship_address_line1', ''),
                    ledger.get('ship_address_line2', ''),
                    ledger.get('ship_address_line3', ''),
                    ledger.get('ship_city', ''),
                    ledger.get('ship_state', ''),
                    ledger.get('ship_pincode', ''),
                    ledger.get('ship_country', ''),
                    ledger.get('gstin', ''),
                    ledger.get('pan', ''),
                    ledger.get('gst_registration_type', ''),
                    ledger.get('tax_rate', ''),
                    ledger.get('bank_name', ''),
                    ledger.get('bank_account_holder', ''),
                    ledger.get('bank_account_number', ''),
                    ledger.get('ifsc_code', ''),
                    ledger.get('branch', ''),
                    ledger.get('swift_code', ''),
                    ledger.get('credit_limit', '0'),
                    ledger.get('credit_period', ''),
                    ledger.get('bill_credit_period', ''),
                    ledger.get('is_revenue', ''),
                    ledger.get('is_deemedpositive', ''),
                    ledger.get('affect_gross_profit', ''),
                    ledger.get('inventory_values_affected', ''),
                    ledger.get('ledger_code', ''),
                    ledger.get('ledger_mobile', ''),
                    ledger.get('ledger_contact', ''),
                    ledger.get('website', ''),
                    ledger.get('nature_of_business', ''),
                    ledger.get('date_of_incorporation', ''),
                    ledger.get('financial_year_from', ''),
                    ledger.get('books_beginning_from', ''),
                    ledger.get('organization_type', ''),
                    ledger.get('company_category', ''),
                    ledger.get('tds_applicable', ''),
                    ledger.get('tcs_applicable', ''),
                    ledger.get('currency', ''),
                    ledger.get('currency_symbol', ''),
                    ledger.get('alterid', ''),
                    ledger.get('reserve1', ''),
                    ledger.get('reserve2', ''),
                    ledger.get('custom_field1', ''),
                    ledger.get('custom_field2', ''),
                ]
                
                for col_idx, value in enumerate(data_row, 1):
                    cell = sheet.cell(row=row_idx, column=col_idx, value=value)
                    cell.alignment = Alignment(vertical='top', wrap_text=False)
            
            # Auto-adjust column widths
            for column in sheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                adjusted_width = min(max_length + 2, 50)
                sheet.column_dimensions[column_letter].width = adjusted_width
            
            # Freeze header row
            sheet.freeze_panes = 'A2'
            
            # Save workbook
            wb.save(filename)
            
            logger.info(f'Exported {len(ledgers)} ledgers to {filename}')
            
            return filename
            
        except Exception as e:
            logger.error(f'Error exporting to Excel: {e}', exc_info=True)
            return None

    def save_ledgers_to_csv(self, ledgers, company_name=""):
        """Legacy CSV export method - kept for backward compatibility"""
        try:
            import csv
            
            if not ledgers:
                logger.warning('Attempted to export empty ledger list to CSV')
                return None
            
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            safe_company_name = "".join(c for c in company_name if c.isalnum() or c in (' ', '-', '_')).strip()
            safe_company_name = safe_company_name[:50]
            
            if safe_company_name:
                filename = f"ledgers_{safe_company_name}_{timestamp}.csv"
            else:
                filename = f"ledgers_{timestamp}.csv"
            
            headers = [
                'Name', 'Parent Group', 'GUID', 'Opening Balance', 'Closing Balance',
                'Address', 'Mailing Name', 'Mobile', 'Email', 'State', 'Country',
                'Pincode', 'GSTIN', 'PAN', 'GST Registration Type',
                'Bank Name', 'Bank Account', 'IFSC Code', 'Branch'
            ]
            
            with open(filename, 'w', newline='', encoding='utf-8-sig') as csvfile:
                writer = csv.writer(csvfile)
                writer.writerow(headers)
                
                for idx, ledger in enumerate(ledgers, 1):
                    row = [
                        ledger.get('name', ''),
                        ledger.get('parent', ''),
                        ledger.get('guid', ''),
                        ledger.get('opening_balance', '0'),
                        ledger.get('closing_balance', '0'),
                        ledger.get('address_line1', ''),
                        ledger.get('mailing_name', ''),
                        ledger.get('mobile', ''),
                        ledger.get('email', ''),
                        ledger.get('state', ''),
                        ledger.get('country', ''),
                        ledger.get('pincode', ''),
                        ledger.get('gstin', ''),
                        ledger.get('pan', ''),
                        ledger.get('gst_registration_type', ''),
                        ledger.get('bank_name', ''),
                        ledger.get('bank_account_number', ''),
                        ledger.get('ifsc_code', ''),
                        ledger.get('branch', '')
                    ]
                    writer.writerow(row)
            
            logger.info(f'Exported {len(ledgers)} ledgers to {filename}')
            
            return filename
            
        except Exception as e:
            logger.error(f'Error exporting to CSV: {e}', exc_info=True)
            return None